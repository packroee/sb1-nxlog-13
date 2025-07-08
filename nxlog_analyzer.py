#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Analyseur de Configuration NXLog

Ce script analyse les fichiers de configuration nxlog et affiche les paramètres 
sous forme de tableau avec descriptions. Il inclut également une fonctionnalité 
de cartographie des flux pour visualiser les routes et connexions entre les sections.

Auteur: Assistant IA
Version: 2.0
"""

import os
import sys
import re
import argparse
import json
from collections import defaultdict, OrderedDict

# Import optionnel de tabulate pour un meilleur affichage
try:
    from tabulate import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False

# Import optionnel d'openpyxl pour Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Dictionnaire des descriptions des paramètres nxlog
PARAMETER_DESCRIPTIONS = {
    # Modules
    'Module': 'Type de module utilisé',
    'ModuleDir': 'Répertoire des modules',
    'CacheDir': 'Répertoire de cache',
    'Pidfile': 'Fichier PID du processus',
    'SpoolDir': 'Répertoire de spool',
    'LogFile': 'Fichier de log principal',
    'LogLevel': 'Niveau de log (DEBUG, INFO, WARNING, ERROR)',
    
    # Fichiers et chemins
    'File': 'Chemin du fichier à traiter',
    'SavePos': 'Sauvegarder la position de lecture',
    'ReadFromLast': 'Lire depuis la fin du fichier',
    'PollInterval': 'Intervalle de polling en secondes',
    'DirCheckInterval': 'Intervalle de vérification du répertoire',
    'ActiveFiles': 'Nombre maximum de fichiers actifs',
    'CloseWhenIdle': 'Fermer quand inactif',
    
    # Réseau
    'Host': 'Adresse IP ou nom d\'hôte',
    'Port': 'Port de connexion',
    'Protocol': 'Protocole utilisé (TCP/UDP)',
    'Listen': 'Adresse d\'écoute',
    'Reconnect': 'Reconnexion automatique',
    'ConnectTimeout': 'Timeout de connexion',
    'AllowUntrusted': 'Autoriser les certificats non fiables',
    
    # SSL/TLS
    'SSL': 'Utiliser SSL/TLS',
    'CertFile': 'Fichier de certificat',
    'CertKeyFile': 'Fichier de clé privée',
    'KeyPass': 'Mot de passe de la clé',
    'CAFile': 'Fichier CA',
    'CADir': 'Répertoire CA',
    'CRLFile': 'Fichier CRL',
    'SSLCompression': 'Compression SSL',
    'SSLCipher': 'Chiffrement SSL',
    'SSLProtocol': 'Version du protocole SSL',
    
    # Format et parsing
    'InputType': 'Type d\'entrée',
    'OutputType': 'Type de sortie',
    'Format': 'Format des données',
    'CSVDelimiter': 'Délimiteur CSV',
    'CSVQuoteChar': 'Caractère de quote CSV',
    'CSVEscapeChar': 'Caractère d\'échappement CSV',
    'CSVQuoteMethod': 'Méthode de quote CSV',
    'Fields': 'Champs à traiter',
    'FieldTypes': 'Types des champs',
    'Delimiter': 'Délimiteur de champs',
    'QuoteChar': 'Caractère de quote',
    'EscapeChar': 'Caractère d\'échappement',
    
    # Performance et buffers
    'BufferSize': 'Taille du buffer',
    'FlushInterval': 'Intervalle de flush',
    'SyncInterval': 'Intervalle de synchronisation',
    'BatchSize': 'Taille des lots',
    'MaxConnections': 'Nombre maximum de connexions',
    'ThreadPoolSize': 'Taille du pool de threads',
    'QueueSize': 'Taille de la queue',
    'HighWaterMark': 'Seuil haut',
    'LowWaterMark': 'Seuil bas',
    
    # Filtres et conditions
    'Condition': 'Condition de filtrage',
    'Priority': 'Priorité de traitement',
    'Exec': 'Code à exécuter',
    'Schedule': 'Planification',
    'First': 'Premier traitement',
    'Every': 'Intervalle de répétition',
    'When': 'Condition temporelle',
    
    # Rotation et archivage
    'CreateDir': 'Créer le répertoire si inexistant',
    'FileMode': 'Permissions du fichier',
    'DirMode': 'Permissions du répertoire',
    'Sync': 'Synchronisation forcée',
    'Truncate': 'Tronquer le fichier',
    'RenameCheck': 'Vérifier le renommage',
    'Recursive': 'Traitement récursif',
    
    # Syslog
    'Facility': 'Facility syslog',
    'Severity': 'Sévérité syslog',
    'Tag': 'Tag syslog',
    'SourceName': 'Nom de la source',
    'ProcessName': 'Nom du processus',
    'Hostname': 'Nom d\'hôte',
    
    # Windows Event Log
    'Channel': 'Canal Windows Event Log',
    'Query': 'Requête XPath',
    'MaxRecords': 'Nombre maximum d\'enregistrements',
    'StartFromFirst': 'Commencer depuis le début',
    'BookmarkXPathFile': 'Fichier bookmark XPath',
    
    # Base de données
    'ConnectionString': 'Chaîne de connexion DB',
    'SQL': 'Requête SQL',
    'Driver': 'Driver de base de données',
    'Username': 'Nom d\'utilisateur DB',
    'Password': 'Mot de passe DB',
    'Table': 'Table de base de données',
    'IdType': 'Type d\'identifiant',
    'CheckInterval': 'Intervalle de vérification',
    
    # HTTP/REST
    'URL': 'URL de destination',
    'HTTPSCertFile': 'Certificat HTTPS',
    'HTTPSKeyFile': 'Clé privée HTTPS',
    'HTTPSCAFile': 'CA HTTPS',
    'ContentType': 'Type de contenu HTTP',
    'AddHeaders': 'En-têtes HTTP additionnels',
    'Compression': 'Compression HTTP',
    
    # Divers
    'User': 'Utilisateur d\'exécution',
    'Group': 'Groupe d\'exécution',
    'NoFreeOnExit': 'Ne pas libérer à la sortie',
    'Locale': 'Paramètres régionaux',
    'DateFormat': 'Format de date',
    'TimeZone': 'Fuseau horaire',
    'Include': 'Fichier à inclure',
    'Define': 'Définition de constante',
    'Extension': 'Extension à charger',
    'FlushLimit': 'Limite de flush',
    'FlushTimeout': 'Timeout de flush',
    'Confirm': 'Confirmation requise',
    'Binary': 'Mode binaire',
    'RawEvent': 'Événement brut',
    'UseUTC': 'Utiliser UTC',
    'PreserveOrder': 'Préserver l\'ordre',
    'IgnoreCase': 'Ignorer la casse',
    'MultiLine': 'Multi-lignes',
    'PatternFile': 'Fichier de patterns',
    'Pattern': 'Pattern de correspondance',
    'Replacement': 'Chaîne de remplacement',
    'Global': 'Remplacement global',
    'CaseSensitive': 'Sensible à la casse',
    'DotAll': 'Mode DotAll',
    'Extended': 'Mode étendu',
    'Multiline': 'Mode multi-lignes',
    'SingleLine': 'Mode ligne unique',
    'Ungreedy': 'Mode non-gourmand'
}

def simple_table_format(data, headers):
    """
    Formate les données en tableau simple sans tabulate
    """
    if not data:
        return "Aucune donnée à afficher"
    
    # Calculer la largeur de chaque colonne
    col_widths = []
    for i, header in enumerate(headers):
        max_width = len(str(header))
        for row in data:
            if i < len(row):
                max_width = max(max_width, len(str(row[i])))
        col_widths.append(max_width)
    
    # Créer la ligne de séparation
    separator = "+" + "+".join("-" * (width + 2) for width in col_widths) + "+"
    
    # Formater l'en-tête
    header_row = "|" + "|".join(f" {str(headers[i]).ljust(col_widths[i])} " for i in range(len(headers))) + "|"
    
    # Formater les données
    result = [separator, header_row, separator]
    
    for row in data:
        formatted_row = "|"
        for i in range(len(headers)):
            cell_value = str(row[i]) if i < len(row) else ""
            formatted_row += f" {cell_value.ljust(col_widths[i])} |"
        result.append(formatted_row)
    
    result.append(separator)
    return "\n".join(result)

def parse_nxlog_config(file_path):
    """
    Parse un fichier de configuration nxlog et extrait les paramètres
    """
    config_data = []
    flow_data = {
        'routes': [],
        'sections': {},
        'flows': []
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {file_path}: {e}")
        return config_data, flow_data
    
    # Supprimer les commentaires (lignes commençant par # ou //)
    lines = []
    for line in content.split('\n'):
        line = line.strip()
        if line and not line.startswith('#') and not line.startswith('//'):
            lines.append(line)
    
    content = '\n'.join(lines)
    
    # Pattern pour identifier les sections
    section_pattern = r'<(\w+)\s+([^>]+)>\s*(.*?)\s*</\1>'
    
    # Trouver toutes les sections
    sections = re.findall(section_pattern, content, re.DOTALL | re.IGNORECASE)
    
    for section_type, section_name, section_content in sections:
        # Nettoyer le nom de section
        section_name = section_name.strip()
        
        # Stocker les informations de section pour la cartographie des flux
        flow_data['sections'][section_name] = {
            'type': section_type,
            'content': section_content.strip()
        }
        
        # Parser les paramètres de la section
        param_pattern = r'(\w+)\s+(.+?)(?=\n\w+\s+|$)'
        params = re.findall(param_pattern, section_content, re.DOTALL)
        
        for param_name, param_value in params:
            param_value = param_value.strip().strip('"\'')
            description = PARAMETER_DESCRIPTIONS.get(param_name, 'Paramètre non documenté')
            
            config_data.append([
                section_type,
                section_name,
                param_name,
                param_value,
                description
            ])
    
    # Parser les routes pour la cartographie des flux
    route_pattern = r'<Route\s+([^>]+)>\s*(.*?)\s*</Route>'
    routes = re.findall(route_pattern, content, re.DOTALL | re.IGNORECASE)
    
    for route_name, route_content in routes:
        route_name = route_name.strip()
        
        # Chercher les définitions de flux dans la route
        path_pattern = r'Path\s+(.+?)(?=\n\w+\s+|$)'
        paths = re.findall(path_pattern, route_content, re.DOTALL)
        
        for path in paths:
            path = path.strip().strip('"\'')
            flow_data['routes'].append({
                'name': route_name,
                'path': path,
                'content': route_content.strip()
            })
    
    # Analyser les flux de données
    analyze_data_flows(flow_data)
    
    return config_data, flow_data

def analyze_data_flows(flow_data):
    """
    Analyse les flux de données basés sur les routes et sections
    """
    flows = []
    
    for route in flow_data['routes']:
        path = route['path']
        route_name = route['name']
        
        # Parser les chemins de type "input1, input2 => processor1 => output1, output2"
        # ou "input1 => output1"
        if '=>' in path:
            steps = [step.strip() for step in path.split('=>')]
            
            # Traiter chaque étape
            for i in range(len(steps) - 1):
                sources = [s.strip() for s in steps[i].split(',')]
                destinations = [d.strip() for d in steps[i + 1].split(',')]
                
                for source in sources:
                    for destination in destinations:
                        if source in flow_data['sections'] and destination in flow_data['sections']:
                            source_info = flow_data['sections'][source]
                            dest_info = flow_data['sections'][destination]
                            
                            flows.append({
                                'route': route_name,
                                'source': source,
                                'source_type': source_info['type'],
                                'source_module': extract_module_from_content(source_info['content']),
                                'destination': destination,
                                'destination_type': dest_info['type'],
                                'destination_module': extract_module_from_content(dest_info['content']),
                                'priority': extract_priority_from_route(route['content']),
                                'condition': extract_condition_from_route(route['content'])
                            })
    
    flow_data['flows'] = flows

def extract_module_from_content(content):
    """
    Extrait le nom du module depuis le contenu d'une section
    """
    module_match = re.search(r'Module\s+(\w+)', content, re.IGNORECASE)
    return module_match.group(1) if module_match else 'N/A'

def extract_priority_from_route(content):
    """
    Extrait la priorité depuis le contenu d'une route
    """
    priority_match = re.search(r'Priority\s+(\d+)', content, re.IGNORECASE)
    return priority_match.group(1) if priority_match else '1'

def extract_condition_from_route(content):
    """
    Extrait la condition depuis le contenu d'une route
    """
    condition_match = re.search(r'Condition\s+(.+?)(?=\n\w+\s+|$)', content, re.DOTALL | re.IGNORECASE)
    return condition_match.group(1).strip().strip('"\'') if condition_match else 'N/A'

def display_config_table(config_data, format_type='table'):
    """
    Affiche les données de configuration dans le format spécifié
    """
    if not config_data:
        print("Aucune configuration trouvée.")
        return
    
    headers = ['Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
    
    if format_type == 'json':
        json_data = []
        for row in config_data:
            json_data.append({
                'section': row[0],
                'section_name': row[1],
                'parameter': row[2],
                'value': row[3],
                'description': row[4]
            })
        print(json.dumps(json_data, indent=2, ensure_ascii=False))
    
    elif format_type == 'csv':
        print(','.join(headers))
        for row in config_data:
            # Échapper les guillemets dans les valeurs CSV
            escaped_row = []
            for cell in row:
                cell_str = str(cell)
                if ',' in cell_str or '"' in cell_str or '\n' in cell_str:
                    cell_str = '"' + cell_str.replace('"', '""') + '"'
                escaped_row.append(cell_str)
            print(','.join(escaped_row))
    
    else:  # format table
        if TABULATE_AVAILABLE:
            print(tabulate(config_data, headers=headers, tablefmt='grid'))
        else:
            print(simple_table_format(config_data, headers))

def display_statistics(config_data):
    """
    Affiche les statistiques de la configuration
    """
    if not config_data:
        print("Aucune donnée pour les statistiques.")
        return
    
    total_params = len(config_data)
    sections = set()
    modules = set()
    section_names = set()
    
    for row in config_data:
        sections.add(row[0])  # Type de section
        section_names.add(row[1])  # Nom de section
        if row[2] == 'Module':  # Si c'est un paramètre Module
            modules.add(row[3])  # Valeur du module
    
    print("=" * 50)
    print("STATISTIQUES DE CONFIGURATION")
    print("=" * 50)
    print(f"Nombre total de paramètres: {total_params}")
    print(f"Nombre de sections: {len(section_names)}")
    print(f"Nombre de modules: {len(modules)}")
    print()
    print(f"Sections trouvées: {', '.join(sorted(sections))}")
    print(f"Modules utilisés: {', '.join(sorted(modules))}")
    print("=" * 50)

def display_flow_mapping(flow_data, config_name="CONFIGURATION"):
    """
    Affiche la cartographie des flux de données
    """
    if not flow_data['flows']:
        print(f"Aucun flux de données trouvé dans {config_name}.")
        return
    
    print("=" * 80)
    print(f"CARTOGRAPHIE DES FLUX - {config_name.upper()}")
    print("=" * 80)
    
    # Statistiques des flux
    total_routes = len(flow_data['routes'])
    total_sections = len(flow_data['sections'])
    total_flows = len(flow_data['flows'])
    
    # Compter les types de sections
    section_types = defaultdict(int)
    for section_info in flow_data['sections'].values():
        section_types[section_info['type']] += 1
    
    # Identifier les sections non connectées
    connected_sections = set()
    for flow in flow_data['flows']:
        connected_sections.add(flow['source'])
        connected_sections.add(flow['destination'])
    
    unconnected_sections = set(flow_data['sections'].keys()) - connected_sections
    
    print("📊 RÉSUMÉ:")
    print(f"  • Routes: {total_routes}")
    print(f"  • Sections: {total_sections}")
    print(f"  • Flux: {total_flows}")
    for section_type, count in sorted(section_types.items()):
        print(f"  • {section_type}s: {count}")
    if unconnected_sections:
        print(f"  • Sections non connectées: {len(unconnected_sections)}")
    print()
    
    if flow_data['flows']:
        print("🔄 FLUX DE DONNÉES:")
        
        # Préparer les données pour le tableau
        flow_table_data = []
        for flow in flow_data['flows']:
            flow_table_data.append([
                flow['route'],
                flow['source'],
                flow['source_type'],
                flow['source_module'],
                '→',
                flow['destination'],
                flow['destination_type'],
                flow['destination_module'],
                flow['priority'],
                flow['condition']
            ])
        
        flow_headers = ['Route', 'Source', 'Type Source', 'Module Source', '', 
                       'Destination', 'Type Dest', 'Module Dest', 'Priorité', 'Condition']
        
        if TABULATE_AVAILABLE:
            print(tabulate(flow_table_data, headers=flow_headers, tablefmt='grid'))
        else:
            print(simple_table_format(flow_table_data, flow_headers))
    
    if unconnected_sections:
        print(f"\n⚠️  SECTIONS NON CONNECTÉES: {', '.join(sorted(unconnected_sections))}")
    
    print("=" * 80)

def create_sample_config():
    """
    Crée un fichier d'exemple de configuration nxlog
    """
    sample_content = """# Configuration NXLog d'exemple
# Fichier généré automatiquement

# Configuration globale
User nxlog
Group nxlog
LogFile /var/log/nxlog/nxlog.log
LogLevel INFO
ModuleDir /usr/lib/nxlog/modules
CacheDir /var/spool/nxlog
Pidfile /var/run/nxlog/nxlog.pid

# Extension pour le format CSV
<Extension csv>
    Module xm_csv
    Fields $timestamp, $hostname, $source, $message
    FieldTypes string, string, string, string
    Delimiter ,
</Extension>

# Input depuis les logs Windows Event Log
<Input eventlog>
    Module im_msvistalog
    Channel System
    SavePos TRUE
    ReadFromLast TRUE
</Input>

# Input depuis un fichier
<Input file>
    Module im_file
    File "/var/log/application.log"
    SavePos TRUE
    ReadFromLast TRUE
    PollInterval 1
</Input>

# Processor pour pattern matching
<Processor pattern>
    Module pm_pattern
    PatternFile /etc/nxlog/patterns.conf
</Processor>

# Output vers syslog UDP
<Output syslog>
    Module om_udp
    Host 192.168.1.100
    Port 514
    Facility local0
    Severity info
</Output>

# Output vers fichier
<Output fileout>
    Module om_file
    File "/var/log/processed.log"
    CreateDir TRUE
    Sync TRUE
</Output>

# Route principale
<Route main>
    Path eventlog, file => pattern => syslog, fileout
    Priority 1
</Route>
"""
    
    filename = "nxlog_sample.conf"
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(sample_content)
        print(f"Fichier d'exemple créé: {filename}")
    except Exception as e:
        print(f"Erreur lors de la création du fichier d'exemple: {e}")

def process_directory(directory_path, stats=False, flows=False, format_type='table'):
    """
    Traite tous les fichiers .conf dans un répertoire
    """
    if not os.path.isdir(directory_path):
        print(f"Erreur: {directory_path} n'est pas un répertoire valide.")
        return {}
    
    config_files = []
    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.conf'):
                config_files.append(os.path.join(root, file))
    
    if not config_files:
        print(f"Aucun fichier .conf trouvé dans {directory_path}")
        return {}
    
    all_configs = {}
    
    for config_file in config_files:
        print(f"\n{'='*60}")
        print(f"ANALYSE DE: {config_file}")
        print(f"{'='*60}")
        
        config_data, flow_data = parse_nxlog_config(config_file)
        all_configs[config_file] = (config_data, flow_data)
        
        if config_data:
            display_config_table(config_data, format_type)
            
            if stats:
                print()
                display_statistics(config_data)
            
            if flows:
                print()
                config_name = os.path.basename(config_file).replace('.conf', '')
                display_flow_mapping(flow_data, config_name)
        else:
            print("Aucune configuration trouvée dans ce fichier.")
    
    return all_configs

def save_to_excel(all_configs, excel_file):
    """
    Sauvegarde toutes les configurations dans un fichier Excel
    """
    if not OPENPYXL_AVAILABLE:
        print("Erreur: openpyxl n'est pas disponible. Installez-le avec: pip install openpyxl")
        return
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Statistiques globales
    stats_ws = wb.create_sheet("Statistiques")
    stats_ws.append(["Fichier", "Nombre de paramètres", "Nombre de sections", "Types de sections"])
    
    for config_file, (config_data, flow_data) in all_configs.items():
        filename = os.path.basename(config_file)
        
        # Feuille de configuration
        config_ws = wb.create_sheet(filename.replace('.conf', ''))
        headers = ['Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
        config_ws.append(headers)
        
        # Appliquer le style d'en-tête
        for cell in config_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Ajouter les données
        for row in config_data:
            config_ws.append(row)
        
        # Ajuster la largeur des colonnes
        for column in config_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            config_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Feuille des flux si disponible
        if flow_data['flows']:
            flow_ws = wb.create_sheet(f"{filename.replace('.conf', '')}_Flux")
            flow_headers = ['Route', 'Source', 'Type Source', 'Module Source', 
                           'Destination', 'Type Dest', 'Module Dest', 'Priorité', 'Condition']
            flow_ws.append(flow_headers)
            
            # Appliquer le style d'en-tête
            for cell in flow_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Ajouter les données de flux
            for flow in flow_data['flows']:
                flow_ws.append([
                    flow['route'], flow['source'], flow['source_type'], flow['source_module'],
                    flow['destination'], flow['destination_type'], flow['destination_module'],
                    flow['priority'], flow['condition']
                ])
            
            # Ajuster la largeur des colonnes
            for column in flow_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                flow_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ajouter aux statistiques
        sections = set()
        for row in config_data:
            sections.add(row[0])
        
        stats_ws.append([
            filename,
            len(config_data),
            len(set(row[1] for row in config_data)),
            ', '.join(sorted(sections))
        ])
    
    # Ajuster les colonnes des statistiques
    for column in stats_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        stats_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Appliquer le style d'en-tête aux statistiques
    for cell in stats_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    try:
        wb.save(excel_file)
        print(f"\nFichier Excel sauvegardé: {excel_file}")
    except Exception as e:
        print(f"Erreur lors de la sauvegarde Excel: {e}")

def save_multiple_csv(all_configs, flows_csv=False):
    """
    Sauvegarde chaque configuration dans un fichier CSV séparé
    """
    for config_file, (config_data, flow_data) in all_configs.items():
        filename = os.path.basename(config_file).replace('.conf', '')
        
        # Fichier CSV de configuration
        csv_filename = f"{filename}_config.csv"
        try:
            with open(csv_filename, 'w', encoding='utf-8') as f:
                headers = ['Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
                f.write(','.join(headers) + '\n')
                
                for row in config_data:
                    escaped_row = []
                    for cell in row:
                        cell_str = str(cell)
                        if ',' in cell_str or '"' in cell_str or '\n' in cell_str:
                            cell_str = '"' + cell_str.replace('"', '""') + '"'
                        escaped_row.append(cell_str)
                    f.write(','.join(escaped_row) + '\n')
            
            print(f"Fichier CSV créé: {csv_filename}")
        except Exception as e:
            print(f"Erreur lors de la création du CSV {csv_filename}: {e}")
        
        # Fichier CSV des flux si demandé
        if flows_csv and flow_data['flows']:
            flow_csv_filename = f"{filename}_flows.csv"
            try:
                with open(flow_csv_filename, 'w', encoding='utf-8') as f:
                    flow_headers = ['Route', 'Source', 'Type Source', 'Module Source', 
                                   'Destination', 'Type Dest', 'Module Dest', 'Priorité', 'Condition']
                    f.write(','.join(flow_headers) + '\n')
                    
                    for flow in flow_data['flows']:
                        row = [flow['route'], flow['source'], flow['source_type'], flow['source_module'],
                               flow['destination'], flow['destination_type'], flow['destination_module'],
                               flow['priority'], flow['condition']]
                        
                        escaped_row = []
                        for cell in row:
                            cell_str = str(cell)
                            if ',' in cell_str or '"' in cell_str or '\n' in cell_str:
                                cell_str = '"' + cell_str.replace('"', '""') + '"'
                            escaped_row.append(cell_str)
                        f.write(','.join(escaped_row) + '\n')
                
                print(f"Fichier CSV des flux créé: {flow_csv_filename}")
            except Exception as e:
                print(f"Erreur lors de la création du CSV des flux {flow_csv_filename}: {e}")

def generate_graphviz_files(all_configs, output_dir="output"):
    """
    Génère les fichiers Graphviz (.dot) pour la visualisation des flux
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Couleurs pour les différents types de sections
    colors = {
        'Input': '#90EE90',      # Vert clair
        'Output': '#FFB6C1',     # Rose clair
        'Processor': '#87CEEB',  # Bleu ciel
        'Extension': '#F0E68C',  # Kaki
        'Route': '#DDA0DD'       # Prune
    }
    
    # Couleurs pour les routes
    route_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8']
    
    synthesis_flows = []  # Pour la cartographie de synthèse
    synthesis_sections = {}
    
def sanitize_node_name(name):
    """
    Nettoie un nom de nœud pour qu'il soit valide en Graphviz
    """
    # Remplacer les caractères problématiques
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', str(name))
    # S'assurer que le nom ne commence pas par un chiffre
    if sanitized and sanitized[0].isdigit():
        sanitized = 'node_' + sanitized
    # S'assurer que le nom n'est pas vide
    if not sanitized:
        sanitized = 'unnamed_node'
    return sanitized

def escape_label(label):
    """
    Échappe les caractères spéciaux dans les labels Graphviz
    """
    if not label:
        return ""
    # Échapper les guillemets et antislashes
    escaped = str(label).replace('\\', '\\\\').replace('"', '\\"')
    # Limiter la longueur pour éviter les labels trop longs
    if len(escaped) > 50:
        escaped = escaped[:47] + "..."
    return escaped

    for config_file, (config_data, flow_data) in all_configs.items():
        filename = os.path.basename(config_file).replace('.conf', '')
        # Nettoyer le nom de fichier
        filename = sanitize_node_name(filename)
        dot_filename = os.path.join(output_dir, f"{filename}_flow.dot")
        
        if not flow_data['flows']:
            continue
        
        # Ajouter à la synthèse
        for section_name, section_info in flow_data['sections'].items():
            prefixed_name = f"{filename}_{sanitize_node_name(section_name)}"
            synthesis_sections[prefixed_name] = {
                'type': section_info['type'],
                'content': section_info['content'],
                'file': filename
            }
        
        for flow in flow_data['flows']:
            synthesis_flows.append({
                'route': f"{filename}_{sanitize_node_name(flow['route'])}",
                'source': f"{filename}_{sanitize_node_name(flow['source'])}",
                'source_type': flow['source_type'],
                'source_module': flow['source_module'],
                'destination': f"{filename}_{sanitize_node_name(flow['destination'])}",
                'destination_type': flow['destination_type'],
                'destination_module': flow['destination_module'],
                'priority': flow['priority'],
                'condition': flow['condition'],
                'file': filename
            })
        
        try:
            with open(dot_filename, 'w', encoding='utf-8') as f:
                f.write('digraph nxlog_flow {\n')
                f.write('    rankdir=LR;\n')
                f.write('    node [shape=box, style=filled];\n')
                f.write('    edge [fontsize=10];\n\n')
                
                # Titre
                f.write(f'    labelloc="t";\n')
                f.write(f'    label="Cartographie des flux NXLog - {escape_label(filename)}";\n\n')
                
                # Définir les nœuds
                for section_name, section_info in flow_data['sections'].items():
                    safe_section_name = sanitize_node_name(section_name)
                    color = colors.get(section_info['type'], '#FFFFFF')
                    module = extract_module_from_content(section_info['content'])
                    label = f"{escape_label(section_name)}\\n({escape_label(section_info['type'])})\\n{escape_label(module)}"
                    f.write(f'    "{safe_section_name}" [fillcolor="{color}", label="{label}"];\n')
                
                f.write('\n')
                
                # Définir les connexions
                route_color_map = {}
                color_index = 0
                
                for flow in flow_data['flows']:
                    safe_source = sanitize_node_name(flow['source'])
                    safe_destination = sanitize_node_name(flow['destination'])
                    
                    if flow['route'] not in route_color_map:
                        route_color_map[flow['route']] = route_colors[color_index % len(route_colors)]
                        color_index += 1
                    
                    edge_color = route_color_map[flow['route']]
                    label = f"Route: {escape_label(flow['route'])}\\nPriorité: {escape_label(flow['priority'])}"
                    if flow['condition'] != 'N/A':
                        condition_short = escape_label(flow['condition'][:20])
                        if len(flow['condition']) > 20:
                            condition_short += "..."
                        label += f"\\nCondition: {condition_short}"
                    
                    f.write(f'    "{safe_source}" -> "{safe_destination}" [color="{edge_color}", label="{label}"];\n')
                
                # Identifier les sections non connectées
                connected_sections = set()
                for flow in flow_data['flows']:
                    connected_sections.add(sanitize_node_name(flow['source']))
                    connected_sections.add(sanitize_node_name(flow['destination']))
                
                all_sections = set(sanitize_node_name(name) for name in flow_data['sections'].keys())
                unconnected_sections = all_sections - connected_sections
                
                if unconnected_sections:
                    f.write('\n    // Sections non connectées\n')
                    for section in unconnected_sections:
                        f.write(f'    "{section}" [style="filled,dashed"];\n')
                
                # Légende
                f.write('\n    // Légende\n')
                f.write('    subgraph cluster_legend {\n')
                f.write('        label="Legende";\n')
                f.write('        style=filled;\n')
                f.write('        fillcolor=lightgray;\n')
                
                legend_items = [
                    ('Input', colors['Input']),
                    ('Output', colors['Output']),
                    ('Processor', colors['Processor']),
                    ('Extension', colors['Extension'])
                ]
                
                for i, (type_name, color) in enumerate(legend_items):
                    f.write(f'        legend_{i} [label="{escape_label(type_name)}", fillcolor="{color}", shape=box];\n')
                
                f.write('    }\n')
                
                f.write('}\n')
            
            print(f"Fichier Graphviz créé: {dot_filename}")
            
            # Créer le script de génération d'images
            script_filename = os.path.join(output_dir, f"{filename}_generate_images.sh")
            with open(script_filename, 'w') as f:
                f.write('#!/bin/bash\n\n')
                f.write(f'# Script de génération d\'images pour {filename}\n\n')
                f.write(f'DOT_FILE="{filename}_flow.dot"\n')
                f.write(f'BASE_NAME="{filename}_flow"\n\n')
                f.write('if ! command -v dot &> /dev/null; then\n')
                f.write('    echo "Erreur: Graphviz n\'est pas installé."\n')
                f.write('    echo "Installez-le avec: sudo apt-get install graphviz (Ubuntu/Debian)"\n')
                f.write('    echo "                  sudo yum install graphviz (CentOS/RHEL)"\n')
                f.write('    echo "                  brew install graphviz (macOS)"\n')
                f.write('    exit 1\n')
                f.write('fi\n\n')
                f.write('echo "Génération des images..."\n\n')
                f.write('# PNG (bitmap haute qualité)\n')
                f.write('dot -Tpng "$DOT_FILE" -o "${BASE_NAME}.png"\n')
                f.write('echo "✓ ${BASE_NAME}.png généré"\n\n')
                f.write('# SVG (vectoriel)\n')
                f.write('dot -Tsvg "$DOT_FILE" -o "${BASE_NAME}.svg"\n')
                f.write('echo "✓ ${BASE_NAME}.svg généré"\n\n')
                f.write('# PDF (imprimable)\n')
                f.write('dot -Tpdf "$DOT_FILE" -o "${BASE_NAME}.pdf"\n')
                f.write('echo "✓ ${BASE_NAME}.pdf généré"\n\n')
                f.write('echo "Génération terminée pour {filename}!"\n')
            
            # Rendre le script exécutable
            os.chmod(script_filename, 0o755)
            print(f"Script de génération créé: {script_filename}")
            
        except Exception as e:
            print(f"Erreur lors de la création du fichier Graphviz {dot_filename}: {e}")
    
    # Générer la cartographie de synthèse
    if synthesis_flows:
        generate_synthesis_graphviz(synthesis_flows, synthesis_sections, output_dir)

def generate_synthesis_graphviz(synthesis_flows, synthesis_sections, output_dir):
    """
    Génère la cartographie de synthèse globale combinant tous les fichiers
    """
    dot_filename = os.path.join(output_dir, "nxlog_synthesis_flow.dot")
    
    # Couleurs pour les différents types de sections
    colors = {
        'Input': '#90EE90',      # Vert clair
        'Output': '#FFB6C1',     # Rose clair
        'Processor': '#87CEEB',  # Bleu ciel
        'Extension': '#F0E68C',  # Kaki
        'Route': '#DDA0DD'       # Prune
    }
    
    # Couleurs de fond pour les clusters (fichiers)
    cluster_colors = ['#F0F8FF', '#F5F5DC', '#F0FFF0', '#FFF8DC', '#F8F8FF', '#F5FFFA']
    
    try:
        with open(dot_filename, 'w', encoding='utf-8') as f:
            f.write('digraph nxlog_synthesis {\n')
            f.write('    rankdir=LR;\n')
            f.write('    node [shape=box, style=filled];\n')
            f.write('    edge [fontsize=8];\n')
            f.write('    compound=true;\n\n')
            
            # Titre
            f.write('    labelloc="t";\n')
            f.write('    label="Cartographie de Synthese NXLog - Vue d ensemble";\n\n')
            
            # Grouper les sections par fichier
            files_sections = defaultdict(list)
            for section_name, section_info in synthesis_sections.items():
                files_sections[section_info['file']].append((section_name, section_info))
            
            # Créer les clusters par fichier
            cluster_index = 0
            for file_name, sections in files_sections.items():
                safe_file_name = sanitize_node_name(file_name)
                cluster_color = cluster_colors[cluster_index % len(cluster_colors)]
                f.write(f'    subgraph cluster_{safe_file_name} {{\n')
                f.write(f'        label="{escape_label(file_name)}.conf";\n')
                f.write('        style=filled;\n')
                f.write(f'        fillcolor="{cluster_color}";\n')
                f.write('        fontsize=12;\n')
                f.write('        fontweight=bold;\n\n')
                
                # Définir les nœuds de ce cluster
                for section_name, section_info in sections:
                    safe_section_name = sanitize_node_name(section_name)
                    color = colors.get(section_info['type'], '#FFFFFF')
                    module = extract_module_from_content(section_info['content'])
                    # Extraire le nom propre de la section (sans le préfixe du fichier)
                    clean_name = section_name
                    if section_name.startswith(f"{file_name}_"):
                        clean_name = section_name[len(f"{file_name}_"):]
                    
                    label = f"{escape_label(clean_name)}\\n({escape_label(section_info['type'])})\\n{escape_label(module)}"
                    f.write(f'        "{safe_section_name}" [fillcolor="{color}", label="{label}"];\n')
                
                f.write('    }\n\n')
                cluster_index += 1
            
            # Définir les connexions
            route_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8']
            route_color_map = {}
            color_index = 0
            
            for flow in synthesis_flows:
                safe_source = sanitize_node_name(flow['source'])
                safe_destination = sanitize_node_name(flow['destination'])
                
                if flow['route'] not in route_color_map:
                    route_color_map[flow['route']] = route_colors[color_index % len(route_colors)]
                    color_index += 1
                
                edge_color = route_color_map[flow['route']]
                # Extraire le nom propre de la route (sans le préfixe du fichier)
                clean_route = flow['route']
                if flow['route'].startswith(f"{flow['file']}_"):
                    clean_route = flow['route'][len(f"{flow['file']}_"):]
                
                label = f"{escape_label(clean_route)}\\nP:{escape_label(flow['priority'])}"
                
                f.write(f'    "{safe_source}" -> "{safe_destination}" [color="{edge_color}", label="{label}"];\n')
            
            # Statistiques de synthèse
            total_files = len(files_sections)
            total_sections = len(synthesis_sections)
            total_flows = len(synthesis_flows)
            
            f.write('\n    // Statistiques de synthèse\n')
            f.write('    subgraph cluster_stats {\n')
            f.write('        label="Statistiques Globales";\n')
            f.write('        style=filled;\n')
            f.write('        fillcolor=lightyellow;\n')
            f.write('        fontsize=10;\n')
            stats_label = f"Fichiers: {total_files}\\nSections: {total_sections}\\nFlux: {total_flows}"
            f.write(f'        stats [shape=note, label="{escape_label(stats_label)}"];\n')
            f.write('    }\n')
            
            # Légende
            f.write('\n    // Légende\n')
            f.write('    subgraph cluster_legend {\n')
            f.write('        label="Legende";\n')
            f.write('        style=filled;\n')
            f.write('        fillcolor=lightgray;\n')
            f.write('        fontsize=10;\n')
            
            legend_items = [
                ('Input', colors['Input']),
                ('Output', colors['Output']),
                ('Processor', colors['Processor']),
                ('Extension', colors['Extension'])
            ]
            
            for i, (type_name, color) in enumerate(legend_items):
                f.write(f'        legend_{i} [label="{escape_label(type_name)}", fillcolor="{color}", shape=box];\n')
            
            f.write('    }\n')
            
            f.write('}\n')
        
        print(f"Cartographie de synthèse créée: {dot_filename}")
        
        # Créer le script de génération d'images pour la synthèse
        script_filename = os.path.join(output_dir, "nxlog_synthesis_generate_images.sh")
        with open(script_filename, 'w') as f:
            f.write('#!/bin/bash\n\n')
            f.write('# Script de génération d\'images pour la cartographie de synthèse\n\n')
            f.write('DOT_FILE="nxlog_synthesis_flow.dot"\n')
            f.write('BASE_NAME="nxlog_synthesis_flow"\n\n')
            f.write('if ! command -v dot &> /dev/null; then\n')
            f.write('    echo "Erreur: Graphviz n\'est pas installé."\n')
            f.write('    echo "Installez-le avec: sudo apt-get install graphviz (Ubuntu/Debian)"\n')
            f.write('    echo "                  sudo yum install graphviz (CentOS/RHEL)"\n')
            f.write('    echo "                  brew install graphviz (macOS)"\n')
            f.write('    exit 1\n')
            f.write('fi\n\n')
            f.write('echo "Génération des images de synthèse..."\n\n')
            f.write('# PNG (bitmap haute qualité)\n')
            f.write('dot -Tpng "$DOT_FILE" -o "${BASE_NAME}.png"\n')
            f.write('echo "✓ ${BASE_NAME}.png généré"\n\n')
            f.write('# SVG (vectoriel)\n')
            f.write('dot -Tsvg "$DOT_FILE" -o "${BASE_NAME}.svg"\n')
            f.write('echo "✓ ${BASE_NAME}.svg généré"\n\n')
            f.write('# PDF (imprimable)\n')
            f.write('dot -Tpdf "$DOT_FILE" -o "${BASE_NAME}.pdf"\n')
            f.write('echo "✓ ${BASE_NAME}.pdf généré"\n\n')
            f.write('echo "Génération de la synthèse terminée!"\n')
        
        # Rendre le script exécutable
        os.chmod(script_filename, 0o755)
        print(f"Script de génération de synthèse créé: {script_filename}")
        
    except Exception as e:
        print(f"Erreur lors de la création de la cartographie de synthèse: {e}")

def main():
    parser = argparse.ArgumentParser(
        description='Analyseur de configuration NXLog avec cartographie des flux',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples d'utilisation:
  %(prog)s nxlog.conf                          # Analyse basique
  %(prog)s nxlog.conf --stats --flows          # Avec statistiques et flux
  %(prog)s --directory /etc/nxlog --flows      # Analyser un répertoire
  %(prog)s --create-sample                     # Créer un exemple
  %(prog)s --directory data --excel rapport.xlsx  # Rapport Excel
  %(prog)s --directory data --graphviz         # Diagrammes Graphviz
        """
    )
    
    parser.add_argument('config_file', nargs='?', help='Fichier de configuration nxlog à analyser')
    parser.add_argument('--create-sample', action='store_true', help='Créer un fichier d\'exemple')
    parser.add_argument('--stats', action='store_true', help='Afficher les statistiques')
    parser.add_argument('--flows', action='store_true', help='Afficher la cartographie des flux')
    parser.add_argument('--format', choices=['table', 'csv', 'json'], default='table', 
                       help='Format de sortie (défaut: table)')
    parser.add_argument('--directory', help='Analyser tous les fichiers .conf dans un répertoire')
    parser.add_argument('--excel-file', help='Sauvegarder dans un fichier Excel')
    parser.add_argument('--csv-multiple', action='store_true', 
                       help='Créer des fichiers CSV séparés pour chaque configuration')
    parser.add_argument('--flows-csv', action='store_true', 
                       help='Inclure les flux dans les fichiers CSV multiples')
    parser.add_argument('--graphviz', action='store_true', 
                       help='Générer les fichiers Graphviz (.dot) pour visualisation')
    
    args = parser.parse_args()
    
    if args.create_sample:
        create_sample_config()
        return
    
    if args.directory:
        all_configs = process_directory(args.directory, args.stats, args.flows, args.format)
        
        if args.excel_file and all_configs:
            save_to_excel(all_configs, args.excel_file)
        
        if args.csv_multiple and all_configs:
            save_multiple_csv(all_configs, args.flows_csv)
        
        if args.graphviz and all_configs:
            generate_graphviz_files(all_configs)
        
        return
    
    if not args.config_file:
        parser.print_help()
        return
    
    if not os.path.exists(args.config_file):
        print(f"Erreur: Le fichier {args.config_file} n'existe pas.")
        return
    
    print(f"Analyse du fichier: {args.config_file}")
    print("=" * 50)
    
    config_data, flow_data = parse_nxlog_config(args.config_file)
    
    if not config_data:
        print("Aucune configuration trouvée dans le fichier.")
        return
    
    display_config_table(config_data, args.format)
    
    if args.stats:
        print()
        display_statistics(config_data)
    
    if args.flows:
        print()
        config_name = os.path.basename(args.config_file).replace('.conf', '')
        display_flow_mapping(flow_data, config_name)

if __name__ == "__main__":
    main()